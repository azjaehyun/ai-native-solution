import argparse
import json
import os
import sys
import base64
from io import BytesIO
from typing import List

# AWS
import boto3

# FAISS
import faiss
import numpy as np

# Word
try:
    from docx import Document
except ImportError:
    Document = None

# PDF
try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

# Excel
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


# 1) chunking 함수들 ---------------------------------------------------------

def dynamic_chunk_text(
    text: str,
    strategy: str = "fixed",
    chunk_size: int = 500,
    overlap_size: int = 50
) -> List[str]:
    """
    일반 텍스트용 동적 청킹 함수.
      - "paragraph": \n\n 기반 문단 분리 + 슬라이딩 윈도우
      - "sentence": 마침표/물음표/느낌표 기반
      - "fixed": 문자 단위 고정 길이
      - "token": (샘플) 공백 단위 n개 단어씩 분할
    """
    text = text.strip()
    if not text:
        return []

    if strategy == "paragraph":
        paragraphs = text.split('\n\n')
        chunks = []
        for para in paragraphs:
            para = para.strip()
            if len(para) <= chunk_size:
                if para:
                    chunks.append(para)
            else:
                start = 0
                while start < len(para):
                    end = start + chunk_size
                    chunk = para[start:end]
                    chunks.append(chunk)
                    start += max(chunk_size - overlap_size, 1)
        return chunks

    elif strategy == "sentence":
        import re
        # 마침표/물음표/느낌표 뒤 공백/줄바꿈을 문장 구분자로 가정
        sentences = re.split(r'(?<=[.!?])\s+', text)
        chunks = []
        for sent in sentences:
            sent = sent.strip()
            if len(sent) <= chunk_size:
                if sent:
                    chunks.append(sent)
            else:
                start = 0
                while start < len(sent):
                    end = start + chunk_size
                    chunk = sent[start:end]
                    chunks.append(chunk)
                    start += max(chunk_size - overlap_size, 1)
        return chunks

    elif strategy == "token":
        # 단순 공백 기준으로 N개 단어씩 쪼개는 예시
        words = text.split()
        chunks = []
        start_idx = 0
        while start_idx < len(words):
            end_idx = start_idx + chunk_size
            chunk_words = words[start_idx:end_idx]
            chunk = " ".join(chunk_words)
            chunks.append(chunk)
            start_idx += max(chunk_size - overlap_size, 1)
        return chunks

    else:
        # "fixed"
        chunks = []
        start = 0
        while start < len(text):
            end = start + chunk_size
            chunk = text[start:end]
            chunks.append(chunk)
            start += chunk_size
        return chunks


def chunk_word(file_bytes: bytes, chunk_size=500, overlap_size=50) -> List[str]:
    """ Word(.docx)을 paragraph 단위로 분할 """
    if Document is None:
        print("docx library not installed.")
        return []

    chunks = []
    try:
        doc = Document(BytesIO(file_bytes))
    except Exception as e:
        print(f"Error loading Word docx: {e}")
        return []

    for i, para in enumerate(doc.paragraphs):
        text = para.text.strip()
        if not text:
            continue
        if len(text) <= chunk_size:
            chunks.append(f"[Paragraph {i+1}] {text}")
        else:
            start = 0
            while start < len(text):
                end = start + chunk_size
                chunk = text[start:end]
                chunks.append(f"[Paragraph {i+1}] {chunk}")
                start += max(chunk_size - overlap_size, 1)

    return chunks


def chunk_pdf(file_bytes: bytes, chunk_size=500, overlap_size=50) -> List[str]:
    """ PDF 파일을 페이지 단위로 텍스트 추출 + 슬라이딩 윈도우 """
    if PyPDF2 is None:
        print("PyPDF2 library not installed.")
        return []

    chunks = []
    try:
        pdf_reader = PyPDF2.PdfReader(BytesIO(file_bytes))
    except Exception as e:
        print(f"Error loading PDF: {e}")
        return []

    for page_index, page in enumerate(pdf_reader.pages):
        try:
            text = page.extract_text() or ""
        except:
            text = ""
        text = text.strip()
        if not text:
            continue

        if len(text) <= chunk_size:
            chunks.append(f"[Page {page_index+1}] {text}")
        else:
            start = 0
            while start < len(text):
                end = start + chunk_size
                chunk = text[start:end]
                chunks.append(f"[Page {page_index+1}] {chunk}")
                start += max(chunk_size - overlap_size, 1)

    return chunks


def chunk_excel(file_bytes: bytes, chunk_size=500, overlap_size=50) -> List[str]:
    """ Excel -> 시트/행 단위 + 슬라이딩 윈도우 """
    if load_workbook is None:
        print("openpyxl library not installed.")
        return []

    chunks = []
    try:
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        print(f"Error loading Excel: {e}")
        return []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            row_str_list = [str(cell) if cell is not None else '' for cell in row]
            row_text = " | ".join(row_str_list).strip()
            if not row_text:
                continue

            if len(row_text) <= chunk_size:
                chunks.append(f"[Sheet: {sheet_name}] {row_text}")
            else:
                start = 0
                while start < len(row_text):
                    end = start + chunk_size
                    chunk = row_text[start:end]
                    chunks.append(f"[Sheet: {sheet_name}] {chunk}")
                    start += max(chunk_size - overlap_size, 1)
    wb.close()
    return chunks


def detect_file_type(filepath: str) -> str:
    """ 파일 확장자로 Word/Excel/PDF/텍스트 구분 (단순 버전) """
    lower_name = filepath.lower()
    if lower_name.endswith(".docx"):
        return "word"
    elif lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
        return "excel"
    elif lower_name.endswith(".pdf"):
        return "pdf"
    else:
        return "text"


def read_and_chunk_file(
    file_path: str,
    chunk_strategy: str,
    chunk_size: int,
    overlap_size: int
) -> List[str]:
    """ 
    로컬 파일을 읽고, 확장자 혹은 전략에 따라 청킹.
    """
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return []

    with open(file_path, "rb") as f:
        file_bytes = f.read()

    file_type = detect_file_type(file_path)
    print(f"Detected file type: {file_type}")

    if file_type == "word":
        return chunk_word(file_bytes, chunk_size, overlap_size)
    elif file_type == "excel":
        return chunk_excel(file_bytes, chunk_size, overlap_size)
    elif file_type == "pdf":
        return chunk_pdf(file_bytes, chunk_size, overlap_size)
    else:
        # text
        text = file_bytes.decode("utf-8", errors="ignore")
        return dynamic_chunk_text(text, strategy=chunk_strategy, 
                                  chunk_size=chunk_size, overlap_size=overlap_size)


# 2) Bedrock 임베딩 + FAISS --------------------------------------------------
def generate_embeddings(chunks: List[str], model_id: str = "amazon.titan-embed-text-v2:0") -> List[np.ndarray]:
    """
    AWS Bedrock를 통해 임베딩을 생성.
    - 로컬에서 실행시, AWS 자격증명(aws configure) 필요
    """
    print(f"Generating embeddings for {len(chunks)} chunks using model {model_id}...")
    bedrock_runtime = boto3.client("bedrock-runtime", region_name="ap-northeast-2")
    embeddings = []

    for i, chunk in enumerate(chunks):
        try:
            cleaned_chunk = chunk.replace('"', '\\"').strip()
            if not cleaned_chunk:
                print(f"[{i+1}/{len(chunks)}] Skipping empty chunk.")
                continue

            payload = {"inputText": cleaned_chunk}
            response = bedrock_runtime.invoke_model(
                modelId=model_id,
                contentType="application/json",
                accept="application/json",
                body=json.dumps(payload)
            )

            response_body = json.loads(response["body"].read())
            embedding_data = response_body.get("embedding")
            if embedding_data is None:
                raise ValueError("No 'embedding' in response")

            embedding = np.array(embedding_data, dtype=np.float32)
            embeddings.append(embedding)
            print(f"[{i+1}/{len(chunks)}] Done. chunk[:30]={chunk[:30]}...")
        except Exception as e:
            print(f"[{i+1}/{len(chunks)}] Error: {e}")
            embeddings.append(None)

    valid_embeddings = [emb for emb in embeddings if emb is not None]
    print(f"Embeddings generation completed. (valid={len(valid_embeddings)}/{len(chunks)})")
    return valid_embeddings


def create_faiss_index(embs: List[np.ndarray]) -> faiss.IndexFlatL2:
    """FAISS 인덱스 생성"""
    if not embs:
        print("No embeddings to create an index.")
        return None
    dimension = embs[0].shape[0]
    index = faiss.IndexFlatL2(dimension)
    index.add(np.vstack(embs))
    print("FAISS index created.")
    return index


def retrieve_relevant_chunks(query: str, faiss_index: faiss.IndexFlatL2, texts: List[str], 
                             model_id: str, top_k: int = 3) -> List[str]:
    """임베딩 → FAISS 검색"""
    if faiss_index is None:
        print("No FAISS index. Can't retrieve.")
        return []

    query_emb = generate_embeddings([query], model_id=model_id)
    if not query_emb:
        return []

    distances, indices = faiss_index.search(np.array([query_emb[0]], dtype=np.float32), top_k)
    result_indices = indices[0]
    print(f"Retrieved indices = {result_indices}, distances = {distances[0]}")
    return [texts[i] for i in result_indices if i < len(texts)]


# 3) 메인 함수 ----------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Local chunking + (optional) embedding test.")
    parser.add_argument("--file", type=str, required=True, help="Local file path to chunk.")
    parser.add_argument("--chunkStrategy", type=str, default="fixed",
                        choices=["fixed", "paragraph", "sentence", "token"],
                        help="How to split text (only for text-based files).")
    parser.add_argument("--maxChunkSize", type=int, default=500, 
                        help="Max chunk size (in characters or 'words' if token strategy).")
    parser.add_argument("--overlapSize", type=int, default=-1,
                        help="Overlap size. If -1, defaults to 10% of maxChunkSize.")
    parser.add_argument("--useEmbedding", action="store_true",
                        help="Whether to generate embeddings via Bedrock and test FAISS retrieval.")
    parser.add_argument("--query", type=str, default="테스트 쿼리",
                        help="Query string for retrieval (only if useEmbedding is enabled).")
    parser.add_argument("--bedrockModel", type=str, default="amazon.titan-embed-text-v2:0",
                        help="Bedrock embedding model ID.")
    parser.add_argument("--topK", type=int, default=3, help="Top K for retrieval.")
    args = parser.parse_args()

    file_path = args.file
    chunk_strategy = args.chunkStrategy
    max_chunk_size = args.maxChunkSize
    overlap_size = args.overlapSize
    if overlap_size < 0:  # -1 이면 자동 설정
        overlap_size = int(max_chunk_size * 0.1)

    print(f"=== Local Chunk Test ===")
    print(f"File = {file_path}")
    print(f"chunkStrategy = {chunk_strategy}")
    print(f"maxChunkSize = {max_chunk_size}")
    print(f"overlapSize = {overlap_size}")
    print(f"useEmbedding = {args.useEmbedding}")
    print("")

    # 1) 청킹
    chunks = read_and_chunk_file(
        file_path=file_path,
        chunk_strategy=chunk_strategy,
        chunk_size=max_chunk_size,
        overlap_size=overlap_size
    )

    print("\n--- Chunks Result ---")
    for i, ch in enumerate(chunks[:10]):
        print(f"Chunk[{i}]: {ch[:100]}{'...' if len(ch)>100 else ''}")
    print(f"Total chunks = {len(chunks)}\n")

    if not args.useEmbedding:
        print("Embedding/FAISS retrieval skipped. Done.")
        return

    # 2) 임베딩 + FAISS 인덱스
    print("Generating embeddings via Bedrock...")
    embeddings = generate_embeddings(chunks, model_id=args.bedrockModel)
    faiss_index = create_faiss_index(embeddings)

    # 3) 검색 테스트
    print(f"\nQuery = {args.query}")
    relevant = retrieve_relevant_chunks(
        query=args.query, 
        faiss_index=faiss_index, 
        texts=chunks,
        model_id=args.bedrockModel,
        top_k=args.topK
    )
    print("\n--- Retrieval Result ---")
    for idx, text in enumerate(relevant, start=1):
        print(f"[{idx}] {text[:200]}{'...' if len(text)>200 else ''}")


if __name__ == "__main__":
    main()



# pip install boto3 faiss-cpu numpy python-docx PyPDF2 openpyxl
# aws configure
# python local_chunk_test.py \
#   --file "/path/to/document.pdf" \
#   --chunkStrategy "sentence" \
#   --maxChunkSize 400

# python local_chunk_test.py \
#   --file "/path/to/document.pdf" \
#   --chunkStrategy "sentence" \
#   --maxChunkSize 400 \
#   --overlapSize 50 \
#   --useEmbedding \
#   --query "이 문서의 핵심 요약은 무엇인가?"


# python local_chunk_test.py \
#   --file "/path/to/long_text.txt" \
#   --chunkStrategy "token" \
#   --maxChunkSize 100 \
#   --overlapSize 10